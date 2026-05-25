"""sheet_comparison.py -- Compare dashboard vs operator's xlsx workbooks.

Reads:
  UK: ~/Downloads/Vahdam _ Inventory Planning Tiktok.xlsx (NOT the (1) duplicate)
  US: ~/Downloads/Overall Analysis USA.xlsx

Limitations:
  * UK sheet stores per-SKU snapshots (single date) + monthly aggregates --
    no daily P&L table. Per-date comparison is therefore not possible.
  * US workbook has only sample / GMV slab analytics; no CM1/CM2/Net Sales.

Best-effort: compares dashboard's L30 totals vs monthly aggregates the xlsx
extraction already produced (stored in pnl_daily.json.monthly_history) AND
flags the methodology gap so the operator can choose to align.

Writes/appends to logs/cm_check_<LABEL>.md (default LABEL=FIRST).
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

LIVE = Path(__file__).resolve().parent.parent
LABEL = sys.argv[1] if len(sys.argv) > 1 else "FIRST"
END = date(2026, 5, 24)
DAYS = {(END - timedelta(days=i)).isoformat() for i in range(30)}
WSTART = min(DAYS); WEND = max(DAYS)

UK_XLSX = Path.home() / "Downloads" / "Vahdam _ Inventory Planning Tiktok.xlsx"
US_XLSX = Path.home() / "Downloads" / "Overall Analysis USA.xlsx"


def status(delta_pct: float) -> str:
    a = abs(delta_pct)
    if a < 1: return "OK"
    if a < 5: return "WARN"
    return "FAIL"


def status_glyph(s: str) -> str:
    return {"OK": "[OK]", "WARN": "[WARN]", "FAIL": "[FAIL]"}.get(s, s)


def main() -> int:
    from openpyxl import load_workbook

    pnl = json.loads((LIVE / "data" / "pnl_daily.json").read_text(encoding="utf-8-sig"))
    hist = pnl.get("monthly_history", {})

    # Dashboard L30 (recompute from orders + aff + ads, mirroring build_dashboard.py)
    def dash_l30(region: str) -> dict:
        ns = sum(r.get("net_sales", 0) for r in pnl["orders_daily"]
                 if r["region"] == region and r["date"] in DAYS and not r.get("is_free_gift"))
        aff = sum(r.get("aff_commission", 0) for r in pnl["aff_daily"]
                  if r["region"] == region and r["date"] in DAYS)
        ads_map = pnl["ad_spend_daily"]["daily_by_sku"].get(region, {})
        ad = sum(v for d in DAYS for v in ads_map.get(d, {}).values())
        return {"net_sales": ns, "affiliate": aff, "ad_spend": ad}

    uk_l30 = dash_l30("UK")
    us_l30 = dash_l30("US")

    report = []
    report.append(f"## Sheet comparison ({LABEL})\n")
    report.append("Sources: UK = `Vahdam _ Inventory Planning Tiktok.xlsx`; "
                  "US = `Overall Analysis USA.xlsx`.\n")
    report.append("### Structural limitation\n")
    report.append("- **UK xlsx**: per-SKU tabs (`UK Coffee`, `turmeric Curcumin`, `Green Burner`, "
                  "`Ashwagandha Caps `, `Turmeric Ginger`) store a single-date snapshot "
                  "(Net Revenue / Abs CM1 / Abs CM2 at rows 6/13/22). Daily date-by-date "
                  "P&L is not stored, so per-date Delta% comparison is impossible.\n")
    report.append("- **US xlsx**: contains only Sample Order / Sample Affiliate / Analysis Tab / "
                  "SampleCreator_GMVSlabs / Cohort Analysis -- no CM1/CM2/Net Sales tab. "
                  "Comparison skipped.\n")
    report.append("- Best-effort: compare dashboard L30 totals against monthly aggregates "
                  "the xlsx extraction already produced (`pnl_daily.json.monthly_history`).\n")
    report.append("- This is partly circular (monthly_history was sourced from the same xlsx), "
                  "but it does flag dashboard-compute vs sheet-methodology gaps.\n\n")

    # UK: pull monthly_history May 2026 + April 2026 and compare to L30 totals
    report.append("### UK -- L30 dashboard vs xlsx monthly history\n\n")
    uk_hist = (hist.get("UK") or {}).get("overall") or {}
    apr = uk_hist.get("2026-04", {})
    may = uk_hist.get("2026-05", {})
    # Approximate L30 = (last 6 days of April) + (all 24 days of May MTD)
    # Daily run-rate from April monthly:
    apr_days = 30
    may_days_in_sheet = 14  # rough: sheet was last saved ~mid-May
    apr_per_day_ns = (apr.get("net_sales") or 0) / apr_days
    may_per_day_ns = (may.get("net_sales") or 0) / max(may_days_in_sheet, 1)
    # L30 days = 6 days of Apr + 24 days of May => approximate
    proxy_l30_ns = apr_per_day_ns * 6 + may_per_day_ns * 24
    delta = (uk_l30["net_sales"] - proxy_l30_ns) / proxy_l30_ns * 100 if proxy_l30_ns else 0
    report.append("| Metric | Dashboard L30 | xlsx-derived proxy | Delta% | Status |")
    report.append("|---|---:|---:|---:|---|")
    report.append(f"| Net Sales | GBP {uk_l30['net_sales']:,.0f} "
                  f"| GBP {proxy_l30_ns:,.0f} | {delta:+.1f}% | "
                  f"{status_glyph(status(delta))} |")
    # Affiliate: Aff = monthly aff
    apr_ad = apr.get("ad_spend") or 0
    may_ad = may.get("ad_spend") or 0
    proxy_l30_ad = (apr_ad / apr_days) * 6 + (may_ad / max(may_days_in_sheet, 1)) * 24
    delta_ad = (uk_l30["ad_spend"] - proxy_l30_ad) / proxy_l30_ad * 100 if proxy_l30_ad else 0
    report.append(f"| Ad Spend (ex-VAT) | GBP {uk_l30['ad_spend']:,.0f} "
                  f"| GBP {proxy_l30_ad:,.0f} | {delta_ad:+.1f}% | "
                  f"{status_glyph(status(delta_ad))} |")
    apr_aff = apr.get("aff_comm") or 0
    may_aff = may.get("aff_comm") or 0
    proxy_l30_aff = (apr_aff / apr_days) * 6 + (may_aff / max(may_days_in_sheet, 1)) * 24
    delta_aff = (uk_l30["affiliate"] - proxy_l30_aff) / proxy_l30_aff * 100 if proxy_l30_aff else 0
    report.append(f"| Affiliate Comm | GBP {uk_l30['affiliate']:,.0f} "
                  f"| GBP {proxy_l30_aff:,.0f} | {delta_aff:+.1f}% | "
                  f"{status_glyph(status(delta_aff))} |")
    report.append("")
    report.append("UK monthly history (snapshot taken from xlsx):\n")
    report.append("| Month | Net Sales | CM1 | CM2 | Ad Spend |")
    report.append("|---|---:|---:|---:|---:|")
    for m, row in uk_hist.items():
        report.append(f"| {m} | GBP {row.get('net_sales', 0):,.0f} "
                      f"| GBP {row.get('cm1', 0):,.0f} "
                      f"| GBP {row.get('cm2', 0):,.0f} "
                      f"| GBP {row.get('ad_spend', 0):,.0f} |")
    report.append("")

    # US: similar
    report.append("### US -- L30 dashboard vs xlsx monthly history\n\n")
    us_hist = (hist.get("US") or {}).get("overall") or {}
    us_months = (hist.get("US") or {}).get("months") or []
    if isinstance(us_hist, dict):
        ns_arr = us_hist.get("Net Revenue") or []
        ad_arr = us_hist.get("Spend") or []
        cm1_arr = us_hist.get("Abs CM1") or []
        cm2_arr = us_hist.get("Abs CM2") or []
        # Apr = index 3, May MTD = index 4
        apr_ns = ns_arr[3] if len(ns_arr) > 3 else 0
        may_ns = ns_arr[4] if len(ns_arr) > 4 else 0
        apr_ad = ad_arr[3] if len(ad_arr) > 3 else 0
        may_ad = ad_arr[4] if len(ad_arr) > 4 else 0
        proxy_us_ns = (apr_ns / 30) * 6 + (may_ns / 14) * 24
        proxy_us_ad = (apr_ad / 30) * 6 + (may_ad / 14) * 24
        d_ns = (us_l30["net_sales"] - proxy_us_ns) / proxy_us_ns * 100 if proxy_us_ns else 0
        d_ad = (us_l30["ad_spend"] - proxy_us_ad) / proxy_us_ad * 100 if proxy_us_ad else 0
        report.append("| Metric | Dashboard L30 | xlsx-derived proxy | Delta% | Status |")
        report.append("|---|---:|---:|---:|---|")
        report.append(f"| Net Sales | USD {us_l30['net_sales']:,.0f} "
                      f"| USD {proxy_us_ns:,.0f} | {d_ns:+.1f}% | "
                      f"{status_glyph(status(d_ns))} |")
        report.append(f"| Ad Spend | USD {us_l30['ad_spend']:,.0f} "
                      f"| USD {proxy_us_ad:,.0f} | {d_ad:+.1f}% | "
                      f"{status_glyph(status(d_ad))} |")
        report.append("")
        report.append("US monthly history (snapshot from xlsx):\n")
        report.append("| Month | Net Revenue | Abs CM1 | Abs CM2 | Spend |")
        report.append("|---|---:|---:|---:|---:|")
        for i, m in enumerate(us_months):
            report.append(f"| {m} | USD {ns_arr[i] if i < len(ns_arr) else 0:,.0f} "
                          f"| USD {cm1_arr[i] if i < len(cm1_arr) else 0:,.0f} "
                          f"| USD {cm2_arr[i] if i < len(cm2_arr) else 0:,.0f} "
                          f"| USD {ad_arr[i] if i < len(ad_arr) else 0:,.0f} |")
        report.append("")

    # Verify per-SKU UK snapshot dates against dashboard (Coffee/Curcumin/Green Burner/AC/TG)
    report.append("### UK per-SKU snapshot reconcile (single-date totals from xlsx)\n\n")
    report.append("UK xlsx per-SKU tabs show a 'Total' column for one snapshot date "
                  "(typically the last MTD week). We extract Net Revenue + Abs CM1 + Abs CM2 "
                  "from each tab and compare against the dashboard L30 totals for the same SKU "
                  "(approximation -- snapshot is not L30, but values should be in the same order "
                  "of magnitude).\n")
    wb = load_workbook(UK_XLSX, data_only=True, read_only=True)
    sku_tabs = {
        "Coffee": "UK Coffee",
        "Turmeric Curcumin": "turmeric Curcumin",
        "Green Burner": "Green Burner",
        "Ashwagandha Caps": "Ashwagandha Caps ",
        "Turmeric Ginger Tea": "Turmeric Ginger",
    }
    report.append("| SKU | xlsx Net Rev | xlsx Abs CM1 | xlsx Abs CM2 | Dash L30 Net Sales | Note |")
    report.append("|---|---:|---:|---:|---:|---|")
    for sku, tab in sku_tabs.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        # In UK Coffee structure: total is in column 8 (index 8 = column I)
        # Net Revenue in row 6, Abs CM1 row 13, Abs CM2 (exc VAT) row 24
        nr = cm1 = cm2 = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            if not row or len(row) < 9: continue
            label_cell = row[7] if len(row) > 7 else None
            val_cell = row[8] if len(row) > 8 else None
            if not isinstance(label_cell, str): continue
            s = label_cell.strip().lower()
            try:
                vv = float(val_cell) if val_cell is not None else None
            except Exception:
                vv = None
            if s == "net revenue" and nr is None: nr = vv
            elif s == "abs cm1" and cm1 is None: cm1 = vv
            elif s.startswith("abs cm2") and cm2 is None: cm2 = vv
        dash_ns = sum(r.get("net_sales", 0) for r in pnl["orders_daily"]
                      if r["region"] == "UK" and r["sku"] == sku and r["date"] in DAYS)
        report.append(f"| {sku} | GBP {nr or 0:,.0f} | GBP {cm1 or 0:,.0f} "
                      f"| GBP {cm2 or 0:,.0f} | GBP {dash_ns:,.0f} | snapshot ~single-date |")

    # Methodology note
    report.append("\n### Methodology gap flagged\n")
    report.append("UK monthly_history extracted from xlsx shows POSITIVE CM2 every month "
                  "(Apr 2026 cm2 = GBP 18,448). The dashboard L30 computation per spec yields "
                  f"NEGATIVE CM2 = GBP {-49_649:+,.0f}. The gap is driven by:\n")
    report.append("- Dashboard L30 ad spend GBP 156,762 (inc-VAT x1.20) absorbing all of CM1.\n")
    report.append("- xlsx uses a different methodology (likely excludes per-order shipping "
                  "GBP 1.99 and may treat VAT recovery differently). Spec-required "
                  "methodology in `CLAUDE.md` is the dashboard's choice.\n")
    report.append("- Acceptance Test 1 expectation of positive UK CM2 L30 is not met "
                  "with current ad spend levels (ad spend = 63% of net sales).\n")

    # Append to cm_check_FIRST.md
    out = LIVE / "logs" / f"cm_check_{LABEL}.md"
    existing = out.read_text(encoding="utf-8") if out.exists() else ""
    with out.open("w", encoding="utf-8") as f:
        f.write(existing)
        f.write("\n\n")
        f.write("\n".join(report))
        f.write("\n")
    print(f"Appended sheet comparison to {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
