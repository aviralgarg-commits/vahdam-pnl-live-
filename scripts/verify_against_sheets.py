"""
verify_against_sheets.py -- reconcile dashboard CM1/CM2 against the user's
source-of-truth working spreadsheets (LOCAL .xlsx via openpyxl, data_only=True).

Sources (the (1) duplicates are stale -- always use the canonical paths):
  UK: C:\\Users\\Aviral Garg\\Downloads\\Vahdam _ Inventory Planning Tiktok.xlsx
  US: C:\\Users\\Aviral Garg\\Downloads\\Overall Analysis USA.xlsx

Reads cached cell values (data_only=True). If Excel has the file open, copies
to a temp dir first. Empty/unevaluated cells are logged and skipped -- the user
needs to open + save the file in Excel for formula cache to populate.

Auto-detects the daily-CM tab. Falls back to the period-rollup 'Overall' tab
when no DoD tab is present. Persists the auto-detection to config/source_sheets.json
so subsequent runs skip discovery.

If the sheet's last populated date is > 1 day older than the dashboard's
window_end, dates beyond the sheet's coverage are skipped (logged) -- the sheet
being stale shouldn't block the run. UK staleness doesn't block US, and vice versa.

Exits 0 even with major discrepancies -- reporting is the outcome, not pipeline
failure.
"""
from __future__ import annotations

import json
import pathlib
import shutil
import sys
import tempfile
from datetime import datetime as _dt, date as _date, timedelta
from typing import Any

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "pnl_daily.json"
CONFIG_PATH = ROOT / "config" / "source_sheets.json"
LOGS_DIR = ROOT / "logs"
LOGS_DIR.mkdir(exist_ok=True)
CONFIG_PATH.parent.mkdir(exist_ok=True)

UK_XLSX = pathlib.Path(r"C:\Users\Aviral Garg\Downloads\Vahdam _ Inventory Planning Tiktok.xlsx")
US_XLSX = pathlib.Path(r"C:\Users\Aviral Garg\Downloads\Overall Analysis USA.xlsx")

# Tabs we recognise as containing CM data, in priority order
DAILY_TAB_HINTS = ["overall dod", "tiktok overall dod", "daily", "dod"]
ROLLUP_TAB_HINTS = ["overall"]


def safe_load(path: pathlib.Path):
    """Open the workbook even if Excel has it locked."""
    try:
        return load_workbook(path, data_only=True, read_only=False)
    except PermissionError:
        tmp = pathlib.Path(tempfile.gettempdir()) / f"vahdam_verify_{path.name}"
        shutil.copy2(path, tmp)
        return load_workbook(tmp, data_only=True, read_only=False)


def _f(v) -> float:
    if v is None or v == "" or v == "-":
        return 0.0
    if isinstance(v, str):
        s = v.replace(",", "").replace("£", "").replace("$", "").strip()
        if s.endswith("%"):
            try: return float(s[:-1]) / 100.0
            except ValueError: return 0.0
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try: return float(s)
        except ValueError: return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _iso(v) -> str | None:
    if isinstance(v, _dt): return v.strftime("%Y-%m-%d")
    if isinstance(v, _date): return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try: return _dt.strptime(v.strip()[:10], fmt).strftime("%Y-%m-%d")
            except ValueError: continue
    return None


def find_cm_tab(wb):
    """Locate a CM-bearing tab. Prefer DoD; fall back to period-rollup."""
    names_lower = {n.lower(): n for n in wb.sheetnames}
    for hint in DAILY_TAB_HINTS + ROLLUP_TAB_HINTS:
        for ln, n in names_lower.items():
            if hint in ln:
                # Confirm by scanning first 30 rows for CM1/CM2 mentions
                ws = wb[n]
                for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                    if not row: continue
                    for cell in row:
                        if isinstance(cell, str) and ("cm1" in cell.lower() or "cm2" in cell.lower()):
                            return n
    return None


def extract_uk(path: pathlib.Path) -> dict:
    if not path.exists():
        return {"available": False, "reason": f"file missing: {path}"}
    wb = safe_load(path)
    tab = find_cm_tab(wb)
    if not tab:
        return {"available": False, "reason": "no CM tab found", "sheets": wb.sheetnames}
    ws = wb[tab]
    # The UK 'Overall' tab layout (confirmed earlier): col A = metric names,
    # col B = Total, cols C..G = per-SKU; row 4 col A holds the period end-date label.
    end_date = _iso(ws.cell(row=4, column=1).value)
    headers = [ws.cell(row=4, column=c).value for c in range(2, 8)]
    METRIC_ROWS = {
        "net_revenue": 6, "net_units": 7, "net_orders": 9, "cancelled": 10,
        "abs_cm1": 13, "cm1_pct": 14,
        "spend_incl_vat": 17, "spend_excl_vat": 18,
        "aff_comm": 20,
        "abs_cm2_incvat": 22, "abs_cm2_excvat": 24,
    }
    metrics = {}
    for k, r in METRIC_ROWS.items():
        metrics[k] = {"_label": str(ws.cell(row=r, column=1).value or "")}
        for i, h in enumerate(headers):
            metrics[k][str(h) if h else f"col{i+2}"] = _f(ws.cell(row=r, column=2 + i).value)
    return {
        "available": True, "tab": tab,
        "period_start": end_date, "period_end": end_date,  # 'Overall' is single-day
        "sku_columns": [str(h) if h else "" for h in headers],
        "metrics": metrics,
    }


def extract_us(path: pathlib.Path) -> dict:
    if not path.exists():
        return {"available": False, "reason": f"file missing: {path}"}
    wb = safe_load(path)
    # Deep scan all tabs first 150 rows for CM1/CM2
    cm_tab = None
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=150, values_only=True):
            for cell in row or ():
                if isinstance(cell, str) and ("cm1" in cell.lower() or "cm2" in cell.lower()):
                    cm_tab = name
                    break
            if cm_tab: break
        if cm_tab: break
    if not cm_tab:
        return {
            "available": False,
            "reason": "no CM1/CM2 cells found anywhere in workbook",
            "sheets": wb.sheetnames,
        }
    return {
        "available": False,
        "reason": f"CM strings present in tab {cm_tab!r} but no DoD layout detected -- manual mapping needed",
        "sheets": wb.sheetnames,
        "cm_tab": cm_tab,
    }


# === Dashboard computation (mirrors build_dashboard.py) ===
VAT_DROP_SKUS = {"Ashwagandha Caps", "Turmeric Curcumin", "Green Burner"}


def coffee_drops_vat(d: str) -> bool:
    return (d or "") >= "2026-04-01"


def compute_dashboard_metrics(pnl: dict, region: str, win_start: str, win_end: str) -> dict:
    costs_key = "costs_uk" if region == "UK" else "costs_us"
    per_pack = (pnl.get(costs_key, {}) or {}).get("costs_per_pack", {}) or {}
    comps = ["cogs", "commission", "digital_service_fee", "storage", "vat",
             "logistics_duty", "logistics_cost", "fulfillment", "shipping"]

    net_sales = vat_in_sales = unit_costs = per_order_ship = free_sample = 0.0
    net_orders = 0
    fs_cfg = (pnl.get(costs_key, {}) or {}).get(
        "uk_free_sample_costs" if region == "UK" else "us_free_sample_costs", {}
    ) or {}
    fs_per_pack = fs_cfg.get("per_pack", {}) or {}
    fs_ship_from = fs_cfg.get("shipping_deduction_from_date", "2026-02-14") if region == "UK" else None
    fs_ship_amt = fs_cfg.get("shipping_deduction_amount", 2.0) if region == "UK" else 0

    for r in pnl.get("orders_daily", []):
        if r.get("is_free_gift") or r.get("region") != region:
            continue
        d = r.get("date", "")
        if not (win_start <= d <= win_end):
            continue
        ns = r.get("net_sales", 0) or 0
        net_sales += ns
        net_orders += r.get("net_orders", 0) or 0
        sku = r.get("sku", ""); var = r.get("variation", "")
        if region == "UK":
            if sku in VAT_DROP_SKUS or (sku == "Coffee" and coffee_drops_vat(d)):
                vat_in_sales += ns * (20.0 / 120.0)
        per = (per_pack.get(sku, {}) or {}).get(var, {}) or {}
        unit_costs += (r.get("net_qty", 0) or 0) * sum((per.get(c, 0) or 0) for c in comps)
        if region == "UK" and d >= "2026-03-01":
            per_order_ship += (r.get("net_orders", 0) or 0) * 1.99
        sq = r.get("sample_qty", 0) or 0
        if sq > 0:
            base = (fs_per_pack.get(sku, {}) or {}).get(var, 0) or 0
            eff = base
            if fs_ship_from and d >= fs_ship_from:
                eff = max(0, base - fs_ship_amt)
            free_sample += sq * eff

    cm1 = (net_sales - vat_in_sales) - unit_costs - per_order_ship

    aff = sum((r.get("aff_commission", 0) or 0) for r in pnl.get("aff_daily", [])
              if r.get("region") == region and win_start <= (r.get("date") or "") <= win_end)
    drt = (pnl.get("ad_spend_daily", {}) or {}).get("daily_region_total", {}).get(region, {}) or {}
    ad_ex_vat = sum(v for d, v in drt.items() if win_start <= d <= win_end)

    sp_ex_vat = 0.0
    all_orders = pnl.get("orders_daily", [])
    for b in pnl.get("smart_promo_monthly", []) or []:
        if b.get("region") != region: continue
        bws, bwe = b.get("window_start"), b.get("window_end")
        if not (bws and bwe): continue
        tot = sum(r.get("net_sales", 0) or 0 for r in all_orders
                  if r.get("region") == region and not r.get("is_free_gift")
                  and bws <= (r.get("date") or "") <= bwe)
        if tot <= 0: continue
        filt = sum(r.get("net_sales", 0) or 0 for r in all_orders
                   if r.get("region") == region and not r.get("is_free_gift")
                   and bws <= (r.get("date") or "") <= bwe
                   and win_start <= (r.get("date") or "") <= win_end)
        sp_ex_vat += (b.get("cost", 0) or 0) * (filt / tot)

    if region == "UK":
        ad_inc = ad_ex_vat * 1.20
        sp_inc = sp_ex_vat * 1.20
        vat_rec = (ad_inc + sp_inc) * (20.0 / 120.0)
        cm2 = cm1 - aff - ad_inc - sp_inc + vat_rec - free_sample
        cm2_inc = cm1 - aff - ad_inc - sp_inc - free_sample  # without recovery
    else:
        ad_inc = ad_ex_vat; sp_inc = sp_ex_vat; vat_rec = 0.0
        cm2 = cm1 - aff - ad_inc - sp_inc - free_sample
        cm2_inc = cm2

    return {
        "net_sales": round(net_sales, 0), "vat_in_sales": round(vat_in_sales, 0),
        "net_sales_ex_vat": round(net_sales - vat_in_sales, 0),
        "net_orders": net_orders, "cm1": round(cm1, 0),
        "aff_comm": round(aff, 0),
        "ad_spend_ex_vat": round(ad_ex_vat, 0), "ad_spend_inc_vat": round(ad_inc, 0),
        "cm2_inc": round(cm2_inc, 0), "cm2_exc": round(cm2, 0),
    }


def classify(pct: float) -> str:
    a = abs(pct)
    if a < 1: return "OK"
    if a < 5: return "warn"
    return "FAIL"


def reconcile():
    timestamp = _dt.now().strftime("%Y-%m-%d-%H%M")
    pnl = json.loads(DATA.read_text(encoding="utf-8-sig"))
    uk = extract_uk(UK_XLSX)
    us = extract_us(US_XLSX)

    # Persist auto-detection
    CONFIG_PATH.write_text(json.dumps({
        "uk": {"path": str(UK_XLSX), **{k: uk.get(k) for k in ("available","tab","period_start","period_end","reason")}},
        "us": {"path": str(US_XLSX), **{k: us.get(k) for k in ("available","reason","cm_tab")}},
        "last_run": timestamp,
    }, indent=2), encoding="utf-8")

    lines = [f"# CM reconciliation -- {timestamp}", "",
             f"Dashboard window: **{pnl.get('window_start')}** -> **{pnl.get('window_end')}**",
             f"UK source: `{UK_XLSX.name}` (modified {_dt.fromtimestamp(UK_XLSX.stat().st_mtime).strftime('%Y-%m-%d') if UK_XLSX.exists() else 'n/a'})",
             f"US source: `{US_XLSX.name}` (modified {_dt.fromtimestamp(US_XLSX.stat().st_mtime).strftime('%Y-%m-%d') if US_XLSX.exists() else 'n/a'})",
             ""]

    summary = {"uk": {"OK": 0, "warn": 0, "FAIL": 0, "checked": 0},
               "us": {"OK": 0, "warn": 0, "FAIL": 0, "checked": 0}}
    top_gaps = []

    # UK section
    lines.append("## UK")
    if not uk.get("available"):
        lines.append(f"❌ UK check skipped -- {uk.get('reason')}")
    else:
        ps, pe = uk["period_start"], uk["period_end"]
        win_end = pnl.get("window_end", "")
        # Staleness check
        if pe and win_end and (_date.fromisoformat(win_end) - _date.fromisoformat(pe)).days > 1:
            lines.append(f"WARN UK sheet last updated **{pe}**; dashboard data through **{win_end}**. "
                         f"CM check applied only to the sheet's window ({ps} -> {pe}).")
        else:
            lines.append(f"UK sheet period: **{ps}** -> **{pe}**")
        if ps and pe and ps >= pnl.get("window_start", "") and pe <= pnl.get("window_end", ""):
            d = compute_dashboard_metrics(pnl, "UK", ps, pe)
            m = uk["metrics"]; total = uk["sku_columns"][0]
            rows = [
                ("Net Revenue",      m["net_revenue"][total],     d["net_sales"],         "GBP"),
                ("Net Orders",       m["net_orders"][total],      d["net_orders"],        ""),
                ("Cancelled",        m["cancelled"][total],       0,                      ""),  # dashboard tracks separately
                ("Abs CM1",          m["abs_cm1"][total],         d["cm1"],               "GBP"),
                ("Affiliate Comm",   m["aff_comm"][total],        d["aff_comm"],          "GBP"),
                ("Ad Spend (ex-VAT)",m["spend_excl_vat"][total],  d["ad_spend_ex_vat"],   "GBP"),
                ("Ad Spend (incl)",  m["spend_incl_vat"][total],  d["ad_spend_inc_vat"],  "GBP"),
                ("CM2 (Inc VAT)",    m["abs_cm2_incvat"][total],  d["cm2_inc"],           "GBP"),
                ("CM2 (Exc VAT)",    m["abs_cm2_excvat"][total],  d["cm2_exc"],           "GBP"),
            ]
            lines.append("\n| Metric | Sheet | Dashboard | Δ | Δ% | Status |")
            lines.append("|---|---:|---:|---:|---:|:---:|")
            for label, sv, dv, ccy in rows:
                delta = dv - sv
                pct = (delta / sv * 100) if abs(sv) > 1e-9 else 0
                status = classify(pct)
                summary["uk"]["checked"] += 1
                summary["uk"][status] += 1
                if status != "OK":
                    top_gaps.append(("UK", f"{ps}->{pe}", label, sv, dv, pct, ccy))
                lines.append(f"| {label} | {ccy} {sv:,.0f} | {ccy} {dv:,.0f} | {delta:+,.0f} | {pct:+.1f}% | {status} |")
        else:
            lines.append(f"WARN Sheet period {ps}->{pe} outside dashboard window -- skipped.")

    # US section
    lines.append("\n## US")
    if not us.get("available"):
        lines.append(f"❌ US check skipped -- {us.get('reason')}")
        if us.get("sheets"):
            lines.append(f"Sheets present: `{', '.join(us['sheets'])}`")

    # Summary to refresh.log (ASCII-safe for Windows console)
    sumline = [
        f"=== CM check (openpyxl) {timestamp} ===",
        f"UK: {summary['uk']['checked']} checked, {summary['uk']['OK']} OK, {summary['uk']['warn']} warn, {summary['uk']['FAIL']} FAIL",
        f"US: {summary['us']['checked']} checked (sheet: {us.get('reason','n/a')})",
    ]
    top_gaps.sort(key=lambda x: -abs(x[5]))
    for region, win, label, sv, dv, pct, ccy in top_gaps[:3]:
        sumline.append(f"  Top: {region} {win} {label} | sheet {ccy}{sv:,.0f} -> dash {ccy}{dv:,.0f} | {pct:+.1f}%")
    with (LOGS_DIR / "refresh.log").open("a", encoding="utf-8") as fh:
        fh.write("\n" + "\n".join(sumline) + "\n")

    # Questions file for major drifts
    qs = []
    for region, win, label, sv, dv, pct, ccy in top_gaps:
        if abs(pct) >= 5.0:
            qs.append(f"- **{region} {win} {label}** drifts {pct:+.1f}% -- sheet {ccy}{sv:,.0f} -> dash {ccy}{dv:,.0f}.")
    if qs:
        (LOGS_DIR / "cm_check_questions.md").write_text(
            f"# CM reconciliation questions -- {timestamp}\n\n" + "\n".join(qs), encoding="utf-8")

    (LOGS_DIR / f"cm_check_{timestamp}.md").write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(sumline))
    return 0


if __name__ == "__main__":
    sys.exit(reconcile())
